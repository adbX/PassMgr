from openpyxl import Workbook
wb = Workbook()

host = raw_input("Enter your host/website")
username = raw_input("Enter your username")
password = raw_input("Enter your password")

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")