from openpyxl import Workbook, load_workbook
import os.path

# check if file exists
if (os.path.isfile("test.xlsx"))==False:
	wb = Workbook()
	ws = wb.active

	# set names of columns
	ws.append(["Host", "Username", "Password"])

else:
	wb = load_workbook("test.xlsx")
	ws = wb.active

# input information
host = raw_input("Enter your host/website")
username = raw_input("Enter your username")
password = raw_input("Enter your password")

ws.append([host, username, password])

wb.save("test.xlsx")
# os.system("pause")