from openpyxl import Workbook, load_workbook
from hashlib import sha256
import pyperclip
import sys
import os


def add_entry():
    # check if file exists
    if os.path.isfile("info.xlsx") == False and os.path.isfile("info.mp4") == False:
        wb = Workbook()
        ws = wb.active

        # set names of columns
        ws.append(["Service", "Username", "Password"])

    elif os.path.isfile("info.mp4") == True:
        mp42xl()
        wb = load_workbook("info.xlsx")
        ws = wb.active

    elif os.path.isfile("info.xlsx") == True:
        wb = load_workbook("info.xlsx")
        ws = wb.active

    else:
        print "File error"

    # input information
    service = raw_input("Enter the service/website: ")
    username = raw_input("Enter your username: ")
    password = raw_input("Enter your password: ")

    ws.append([service, username, password])
    wb.save("info.xlsx")
    xl2mp4()


def generate():
    alphabet = ('abcdefghijklmnopqrstuvwxyz'
                'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
                '0123456789!@#$%^&*()-_')

    my_key = '3141592653589adb3141592653589'

    def create_hex(salt, plaintext):
        return sha256(salt + plaintext).hexdigest()

    def make_password(plaintext, service):
        salt = create_hex(my_key, service)[:20]
        hsh = create_hex(salt, plaintext)
        return ''.join((salt, hsh))

    def password(plaintext, service, length=10, alpha=alphabet):
        raw_hex = make_password(plaintext, service)

        # Convert the hex to dec
        num = int(raw_hex, 16)

        num_chars = len(alpha)

        chars = []
        while len(chars) < length:
            num, index = divmod(num, num_chars)
            chars.append(alpha[index])

        clip = ''.join(chars)
        print clip
        pyperclip.copy(clip)
        print "Copied to Clipboard!"

    master_pass = raw_input("Enter your master password: ")
    service = raw_input("Enter the service/website: ")
    password(master_pass, service)


def view_entries():
    mp42xl()
    os.system("start "+"info.xlsx")

    close_inp = raw_input("Enter 'x' if you closed the file")
    if close_inp == "x":
        xl2mp4()


def purge():
    f = open("info.mp4", "wb")
    f.write("*"*os.path.getsize("info.mp4"))
    f.close()
    os.unlink("info.mp4")
    f.close()


def xl2mp4():
    x = list(os.path.splitext("info.xlsx"))
    x[1] = ".mp4"
    y = "".join(x)
    os.rename("info.xlsx", y)


def mp42xl():
    x = list(os.path.splitext("info.mp4"))
    x[1] = ".xlsx"
    y = "".join(x)
    os.rename("info.mp4", y)


while True:
    print "Welcome to PassMgr!"
    print "1 for Entering new information"
    print "2 for Generating a password"
    print "3 for Purging all information"
    print "4 for Viewing existing information"
    print "5 to Exit"

    option = int(raw_input("Enter one of the above options: "))
    if option == 1:
        add_entry()
    elif option == 2:
        generate()
    elif option == 3:
        purge()
    elif option == 4:
        view_entries()
    elif option == 5:
        sys.exit()
