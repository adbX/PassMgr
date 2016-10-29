from openpyxl import Workbook, load_workbook
import os.path

def AddEntry():
	# check if file exists
	if (os.path.isfile("test.xlsx"))==False:
		wb = Workbook()
		ws = wb.active

		# set names of columns
		ws.append(["Service", "Username", "Password"])

	else:
		wb = load_workbook("test.xlsx")
		ws = wb.active

	# input information
	service = raw_input("Enter the service/website: ")
	username = raw_input("Enter your username: ")
	password = raw_input("Enter your password: ")

	ws.append([service, username, password])

	wb.save("test.xlsx")